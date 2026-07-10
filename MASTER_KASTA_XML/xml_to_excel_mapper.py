import argparse
import os
import re
import requests
import lxml.etree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_args():
    parser = argparse.ArgumentParser(description="Convert any XML price feed to Excel for mapping.")
    parser.add_argument("--input", required=True, help="Path to local XML file or URL to XML feed")
    parser.add_argument("--output", required=True, help="Path to output Excel file (.xlsx)")
    return parser.parse_args()

def load_xml(source):
    if source.startswith("http://") or source.startswith("https://"):
        print(f"Downloading XML from URL: {source} ...")
        r = requests.get(source, timeout=60)
        r.raise_for_status()
        # Parse from string bytes to avoid encoding mismatches
        return ET.fromstring(r.content)
    else:
        print(f"Loading local XML file: {source} ...")
        if not os.path.exists(source):
            raise FileNotFoundError(f"Local file not found: {source}")
        tree = ET.parse(source)
        return tree.getroot()

def build_category_map(root):
    category_map = {}
    # Find all category tags (YML standard)
    for cat in root.xpath("//category"):
        cat_id = cat.get("id")
        if cat_id:
            category_map[cat_id] = {
                "name": (cat.text or "").strip(),
                "parentId": cat.get("parentId") or cat.get("parent_id") or cat.get("parentID") or ""
            }
    return category_map

def extract_offers(root):
    # Check if we have YML offers
    offers = root.xpath("//offer")
    if not offers:
        # Check if we have RSS items
        offers = root.xpath("//item")
    if not offers:
        # Generic fallback: any elements that look like products (e.g. elements inside a root)
        # We can look for tags that are commonly used
        offers = root.xpath("//product")
    return offers

def clean_text(text):
    if not text:
        return ""
    # Strip basic control characters
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', str(text))
    return text.strip()

def process_offer(offer, category_map):
    # 1. Offer ID
    offer_id = offer.get("id") or offer.findtext("id") or offer.findtext("{http://base.google.com/ns/1.0}id") or ""
    offer_id = clean_text(offer_id)
    
    # Extract prefix if ID is formatted like 1000_1234
    prefix = ""
    if "_" in offer_id:
        parts = offer_id.split("_")
        if parts[0].isdigit():
            prefix = parts[0]

    # 2. Name
    name = (offer.findtext("name_ua") or 
            offer.findtext("name") or 
            offer.findtext("title") or 
            offer.findtext("{http://base.google.com/ns/1.0}title") or "")
    name = clean_text(name)
    
    # 3. Description
    desc = (offer.findtext("description_ua") or 
            offer.findtext("description") or 
            offer.findtext("{http://base.google.com/ns/1.0}description") or "")
    desc = clean_text(desc)
    
    # 4. Price and Old Price
    price = offer.findtext("price") or offer.findtext("{http://base.google.com/ns/1.0}price") or "0"
    price_old = offer.findtext("price_old") or offer.findtext("oldprice") or offer.findtext("{http://base.google.com/ns/1.0}sale_price") or ""
    
    price = clean_text(price)
    price_old = clean_text(price_old)
    
    # 5. Quantity and Availability
    qty = offer.findtext("stock_quantity") or offer.findtext("quantity") or offer.findtext("{http://base.google.com/ns/1.0}quantity") or ""
    qty = clean_text(qty)
    
    available = offer.get("available") or offer.findtext("available") or offer.findtext("{http://base.google.com/ns/1.0}availability") or "true"
    available = clean_text(available)
    
    # 6. Vendor
    vendor = offer.findtext("vendor") or offer.findtext("brand") or offer.findtext("{http://base.google.com/ns/1.0}brand") or ""
    vendor = clean_text(vendor)
    
    # 7. Category
    cat_id = offer.findtext("categoryId") or offer.findtext("category_id") or offer.findtext("{http://base.google.com/ns/1.0}product_type") or ""
    cat_id = clean_text(cat_id)
    cat_name = ""
    if cat_id and cat_id in category_map:
        cat_name = category_map[cat_id]["name"]
        
    # 8. Pictures
    pics = []
    # Check YML standard <picture> tags
    for p in offer.findall("picture"):
        if p.text:
            pics.append(clean_text(p.text))
    # Check Google XML standard <g:image_link> and <g:additional_image_link>
    g_img = offer.findtext("{http://base.google.com/ns/1.0}image_link")
    if g_img:
        pics.append(clean_text(g_img))
    for add_img in offer.findall("{http://base.google.com/ns/1.0}additional_image_link"):
        if add_img.text:
            pics.append(clean_text(add_img.text))
            
    pictures_str = ",".join(pics)
    
    # 9. Parameters (Color, Size, and Others)
    color_val = ""
    size_val = ""
    other_params = []
    
    # YML standard parameters
    for param in offer.findall("param"):
        p_name = (param.get("name") or "").strip()
        p_val = (param.text or "").strip()
        if not p_name:
            continue
            
        p_name_low = p_name.lower()
        if p_name_low in ("колір", "цвет", "color"):
            color_val = p_val
        elif any(w in p_name_low for w in ("розмір", "размер", "size")):
            size_val = p_val
        else:
            other_params.append(f"{p_name}: {p_val}")
            
    # Google standard attributes
    g_color = offer.findtext("{http://base.google.com/ns/1.0}color")
    if g_color:
        color_val = clean_text(g_color)
    g_size = offer.findtext("{http://base.google.com/ns/1.0}size")
    if g_size:
        size_val = clean_text(g_size)
        
    other_params_str = "; ".join(other_params)
    
    return {
        "Offer ID": offer_id,
        "Prefix": prefix,
        "Назва UA": name,
        "Опис": desc,
        "Ціна": price,
        "Стара ціна": price_old,
        "Кількість": qty,
        "Наявність": available,
        "Бренд": vendor,
        "ID Поточної Категорії": cat_id,
        "Назва Поточної Категорії": cat_name,
        "ID НОВОЇ Категорії": "",     # Empty for mapping
        "Назва НОВОЇ Категорії": "",   # Empty for mapping
        "Посилання на фото": pictures_str,
        "Колір": color_val,
        "Розмір": size_val,
        "Інші Характеристики": other_params_str
    }

def main():
    args = parse_args()
    
    try:
        root = load_xml(args.input)
    except Exception as e:
        print(f"Error loading XML: {e}")
        return
        
    category_map = build_category_map(root)
    print(f"Loaded {len(category_map)} categories from XML.")
    
    offers = extract_offers(root)
    print(f"Found {len(offers)} offers/items in XML.")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Mapping"
    
    headers = [
        "Offer ID", "Prefix", "Назва UA", "Опис", "Ціна", "Стара ціна", 
        "Кількість", "Наявність", "Бренд", "ID Поточної Категорії", 
        "Назва Поточної Категорії", "ID НОВОЇ Категорії", "Назва НОВОЇ Категорії", 
        "Посилання на фото", "Колір", "Розмір", "Інші Характеристики"
    ]
    
    # Styling variables
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Write headers
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Write data
    row_count = 0
    for offer in offers:
        row_data = process_offer(offer, category_map)
        row_list = [row_data[h] for h in headers]
        ws.append(row_list)
        row_count += 1
        
        # Apply basic cell borders and alignments
        current_row = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            if headers[col_num - 1] in ("Offer ID", "Prefix", "Ціна", "Стара ціна", "Кількість", "Наявність", "ID Поточної Категорії", "ID НОВОЇ Категорії", "Колір", "Розмір"):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
        if row_count % 1000 == 0:
            print(f"Processed {row_count} items...")
            
    # Set reasonable column widths
    ws.column_dimensions["A"].width = 15  # Offer ID
    ws.column_dimensions["B"].width = 10  # Prefix
    ws.column_dimensions["C"].width = 45  # Назва
    ws.column_dimensions["D"].width = 50  # Опис
    ws.column_dimensions["E"].width = 10  # Ціна
    ws.column_dimensions["F"].width = 12  # Стара ціна
    ws.column_dimensions["G"].width = 10  # Кількість
    ws.column_dimensions["H"].width = 12  # Наявність
    ws.column_dimensions["I"].width = 15  # Бренд
    ws.column_dimensions["J"].width = 22  # ID поточної кат
    ws.column_dimensions["K"].width = 28  # Назва поточної кат
    ws.column_dimensions["L"].width = 22  # ID нової кат (Highlight fill for mapping)
    ws.column_dimensions["M"].width = 28  # Назва нової кат
    ws.column_dimensions["N"].width = 30  # Посилання на фото
    ws.column_dimensions["O"].width = 15  # Колір
    ws.column_dimensions["P"].width = 15  # Розмір
    ws.column_dimensions["Q"].width = 40  # Інші характеристики
    
    # Highlight new mapping columns to make them pop out
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=12).fill = highlight_fill
        ws.cell(row=r, column=13).fill = highlight_fill
        
    print(f"Saving Excel sheet to: {args.output} ...")
    wb.save(args.output)
    print("Done! Export completed successfully.")

if __name__ == "__main__":
    main()
