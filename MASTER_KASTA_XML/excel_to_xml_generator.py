import argparse
import os
import re
import json
import lxml.etree as ET
import openpyxl
from datetime import datetime

SYNONYMS = {
    "Offer ID": ["offer id", "id", "артикул", "код", "код товару", "sku", "идентификатор"],
    "Prefix": ["prefix", "префікс", "префикс"],
    "Назва UA": ["назва ua", "назва", "name", "title", "название", "назва укр"],
    "Опис": ["опис", "опис ua", "description", "описание"],
    "Ціна": ["ціна", "price", "цена", "вартість"],
    "Стара ціна": ["стара ціна", "price_old", "oldprice", "старая цена"],
    "Кількість": ["кількість", "stock_quantity", "quantity", "кол-во", "количество", "qty"],
    "Наявність": ["наявність", "available", "наличие"],
    "Бренд": ["бренд", "vendor", "brand", "виробник", "производитель"],
    "ID Поточної Категорії": ["id поточної категорії", "id категорії", "category id", "category_id", "категория id"],
    "Назва Поточної Категорії": ["назва поточної категорії", "назва категорії", "category name", "category_name", "категория назва"],
    "ID НОВОЇ Категорії": ["id нової категорії", "новий id категорії", "new category id", "new_category_id", "нова категорія id"],
    "Назва НОВОЇ Категорії": ["назва нової категорії", "нова назва категорії", "new category name", "new_category_name"],
    "Посилання на фото": ["посилання на фото", "картинка", "фото", "picture", "pictures", "images", "image_link"],
    "Колір": ["колір", "color", "цвет"],
    "Розмір": ["розмір", "size", "размер"],
    "Інші Характеристики": ["інші характеристики", "характеристики", "params", "parameters", "параметры"]
}

def parse_args():
    parser = argparse.ArgumentParser(description="Convert an edited Excel sheet back to YML XML format.")
    parser.add_argument("--input", required=True, help="Path to input Excel file (.xlsx)")
    parser.add_argument("--output", required=True, help="Path to output XML file (.xml)")
    return parser.parse_args()

def resolve_columns(headers):
    resolved_headers = {}
    clean_headers = [(str(h).strip().lower(), idx) for idx, h in enumerate(headers) if h is not None]
    
    for standard_name, synonyms in SYNONYMS.items():
        standard_lower = standard_name.lower()
        matched_idx = None
        
        # Exact match check
        for h_lower, idx in clean_headers:
            if h_lower == standard_lower:
                matched_idx = idx
                break
                
        # Synonym check
        if matched_idx is None:
            for syn in synonyms:
                for h_lower, idx in clean_headers:
                    if h_lower == syn:
                        matched_idx = idx
                        break
                if matched_idx is not None:
                    break
                    
        if matched_idx is not None:
            resolved_headers[standard_name] = matched_idx
            
    return resolved_headers

def sanitize_price(val):
    if not val:
        return ""
    val = re.sub(r'\s+', '', str(val))
    val = re.sub(r'(грн|грн\.|\$|€|руб|uah|usd|eur)', '', val, flags=re.IGNORECASE)
    val = val.replace(",", ".")
    match = re.search(r'^\d+(\.\d+)?', val)
    if match:
        return match.group(0)
    return ""

def sanitize_qty(val):
    if not val:
        return ""
    val = re.sub(r'\s+', '', str(val))
    match = re.search(r'^\d+', val)
    if match:
        return match.group(0)
    return ""

def parse_other_params(params_str):
    params = []
    if not params_str:
        return params
    parts = params_str.split("; ")
    for part in parts:
        if ":" in part:
            name, val = part.split(":", 1)
            params.append((name.strip(), val.strip()))
    return params

def main():
    args = parse_args()
    
    print(f"Loading Excel file: {args.input} ...")
    if not os.path.exists(args.input):
        print(f"Error: Excel file not found at {args.input}")
        return
        
    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    header_map = resolve_columns(headers)
    
    # Verify minimal headers are present (Fuzzy matched)
    required_standards = ["Offer ID", "Назва UA", "Ціна"]
    for req in required_standards:
        if req not in header_map:
            print(f"Error: Could not resolve column for '{req}' in Excel file.")
            print(f"Headers found: {headers}")
            return

    def get_val(row, col_name):
        idx = header_map.get(col_name)
        if idx is not None:
            val = row[idx].value
            return str(val).strip() if val is not None else ""
        return ""

    # Setup category memory JSON
    rules_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "category_mapping_rules.json")
    mapping_rules = {}
    if os.path.exists(rules_path):
        try:
            with open(rules_path, "r", encoding="utf-8") as f:
                mapping_rules = json.load(f)
        except Exception:
            pass

    # Build XML tree
    yml_catalog = ET.Element("yml_catalog", date=datetime.now().strftime("%Y-%m-%d %H:%M"))
    shop = ET.SubElement(yml_catalog, "shop")
    
    ET.SubElement(shop, "name").text = "AVI KASTA"
    ET.SubElement(shop, "url").text = "https://avi.in.ua"
    
    currencies = ET.SubElement(shop, "currencies")
    ET.SubElement(currencies, "currency", id="UAH", rate="1")
    
    categories = ET.SubElement(shop, "categories")
    offers = ET.SubElement(shop, "offers")
    
    unique_categories = {}
    
    # Store dynamic mapping updates
    mapping_updates_count = 0
    row_count = 0
    
    rows = list(ws.iter_rows())[1:]
    for row in rows:
        offer_id = get_val(row, "Offer ID")
        if not offer_id:
            continue
            
        name_ua = get_val(row, "Назва UA")
        desc = get_val(row, "Опис")
        price = sanitize_price(get_val(row, "Ціна"))
        price_old = sanitize_price(get_val(row, "Стара ціна"))
        qty = sanitize_qty(get_val(row, "Кількість"))
        available = get_val(row, "Наявність") or "true"
        vendor = get_val(row, "Бренд")
        
        # Category Resolution & Memory Tracking
        current_cat_id = get_val(row, "ID Поточної Категорії")
        current_cat_name = get_val(row, "Назва Поточної Категорії")
        new_cat_id = get_val(row, "ID НОВОЇ Категорії")
        new_cat_name = get_val(row, "Назва НОВОЇ Категорії")
        
        if new_cat_id:
            cat_id = new_cat_id
            cat_name = new_cat_name or f"Категорія {new_cat_id}"
            
            # Save to memory if it's a new rule or changed
            if current_cat_id and (current_cat_id not in mapping_rules or mapping_rules[current_cat_id].get("new_id") != new_cat_id):
                mapping_rules[current_cat_id] = {
                    "new_id": new_cat_id,
                    "new_name": new_cat_name
                }
                mapping_updates_count += 1
        else:
            cat_id = current_cat_id
            cat_name = current_cat_name or f"Категорія {current_cat_id}"
            
        if cat_id:
            unique_categories[cat_id] = cat_name
            
        pictures_str = get_val(row, "Посилання на фото")
        color = get_val(row, "Колір")
        size = get_val(row, "Розмір")
        other_params_str = get_val(row, "Інші Характеристики")
        
        # Variant Splitter: Check if size has multiple values
        sizes = []
        if size:
            # Separators: comma, semicolon, slash
            s_clean = size.replace(";", ",").replace("/", ",")
            sizes = [item.strip() for item in s_clean.split(",") if item.strip()]
            
        # Helper to generate a single XML offer
        def build_xml_offer(o_id, sz_val, group_id_val=None):
            offer_el = ET.SubElement(offers, "offer", id=o_id, available=available)
            if group_id_val:
                ET.SubElement(offer_el, "group_id").text = group_id_val
                
            ET.SubElement(offer_el, "price").text = price
            if price_old:
                ET.SubElement(offer_el, "price_old").text = price_old
                
            if qty:
                ET.SubElement(offer_el, "stock_quantity").text = qty
                
            ET.SubElement(offer_el, "currencyId").text = "UAH"
            ET.SubElement(offer_el, "categoryId").text = cat_id
            
            if pictures_str:
                for pic_url in pictures_str.split(","):
                    pic_url = pic_url.strip()
                    if pic_url:
                        ET.SubElement(offer_el, "picture").text = pic_url
                        
            if vendor:
                ET.SubElement(offer_el, "vendor").text = vendor
                
            ET.SubElement(offer_el, "name_ua").text = name_ua
            
            if desc:
                ET.SubElement(offer_el, "description_ua").text = ET.CDATA(desc)
                
            if color:
                ET.SubElement(offer_el, "param", name="Колір").text = color
            if sz_val:
                ET.SubElement(offer_el, "param", name="Розмір").text = sz_val
                
            # Other params
            other_params = parse_other_params(other_params_str)
            for p_name, p_val in other_params:
                ET.SubElement(offer_el, "param", name=p_name).text = p_val
                
        # Split or write single
        if len(sizes) > 1:
            for sz in sizes:
                sz_slug = re.sub(r'[^a-zA-Z0-9]', '', sz).lower()
                variant_id = f"{offer_id}_{sz_slug}"
                build_xml_offer(variant_id, sz, offer_id)
                row_count += 1
        else:
            sz_val = sizes[0] if sizes else size
            build_xml_offer(offer_id, sz_val)
            row_count += 1
            
        if row_count % 1000 == 0:
            print(f"Constructed {row_count} XML offers...")
            
    # Write categories block
    for cid, cname in unique_categories.items():
        ET.SubElement(categories, "category", id=cid).text = cname
        
    # Save category memory updates
    if mapping_updates_count > 0:
        try:
            with open(rules_path, "w", encoding="utf-8") as f:
                json.dump(mapping_rules, f, ensure_ascii=False, indent=2)
            print(f"Saved {mapping_updates_count} new category mapping rules to memory JSON.")
        except Exception as e:
            print(f"Failed to write mapping memory to JSON: {e}")
            
    print(f"Saving final XML to: {args.output} ...")
    tree = ET.ElementTree(yml_catalog)
    with open(args.output, "wb") as f:
        tree.write(f, xml_declaration=True, encoding="UTF-8", pretty_print=True)
        
    print(f"Done! Recreated XML with {row_count} offers and {len(unique_categories)} categories.")

if __name__ == "__main__":
    main()
